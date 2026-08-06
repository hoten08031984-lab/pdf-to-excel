import os
import sys
import re
import datetime
import fitz  # PyMuPDF
from rapidocr_onnxruntime import RapidOCR
import openpyxl
from openpyxl.styles import Alignment, Font, PatternFill

sys.stdout.reconfigure(encoding='utf-8')

ocr = RapidOCR()

BASE_DIR = os.path.dirname(os.path.abspath(__file__))
EXCEL_PATH = os.path.join(BASE_DIR, 'Xăng.xlsx')
PDF_DIR = os.path.join(BASE_DIR, 'Hóa đơn')

# Styles
CENTER_ALIGN = Alignment(horizontal='center', vertical='center')
FMT_CURRENCY = '_-* #,##0_-;\\-* #,##0_-;_-* "-"_-;_-@_-'
FMT_QTY = '#,##0.000'

# Font styles for visual distinction
FONT_RED = Font(color='FF0000')  # Red text for old data
NO_FILL = PatternFill(fill_type=None)  # Clear any custom background (restore table default)

def parse_pdf_invoice(pdf_path):
    """Extract invoice details using PyMuPDF + RapidOCR."""
    with open(pdf_path, 'rb') as f:
        pdf_bytes = f.read()
    doc = fitz.open(stream=pdf_bytes, filetype='pdf')
    page = doc[0]
    pix = page.get_pixmap(dpi=200)
    img_bytes = pix.tobytes('png')
    
    result, _ = ocr(img_bytes)
    if not result:
        return None
        
    lines = [item[1].strip() for item in result]
    full_text = "\n".join(lines)
    
    # 1. Invoice Number (Số hóa đơn)
    inv_num = None
    # Exclude lines containing bank accounts or tax codes
    clean_text = "\n".join([line for line in full_text.split("\n") if not re.search(r'Acc\s*No|t[àa]i khoản|Tax\s*Code|mã số thuế', line, re.IGNORECASE)])
    
    m_inv = re.search(r'\(No\.?\):?\s*(\d{1,8})', clean_text, re.IGNORECASE)
    if not m_inv:
        m_inv = re.search(r'(?:S[ốóe60ô]|So|No|S6|S0)[:\s\.]*(?:\(?No\.?\)?[:\s\.]*)?\n?\s*(\d{1,8})', clean_text, re.IGNORECASE)
    if not m_inv:
        m_inv = re.search(r'(?:S[ốóe60ô]|So|S6|S0)[:.\s]+\n?\s*(\d{1,8})', clean_text, re.IGNORECASE)
    
    if m_inv:
        inv_num = int(m_inv.group(1))
    else:
        # Fallback: find all digit groups in filename, pick the last group with length >= 4 (e.g. 164323 in HDGTGT_K26TXT_164323.pdf)
        digit_groups = re.findall(r'\d+', os.path.basename(pdf_path))
        valid_groups = [g for g in digit_groups if len(g) >= 4]
        if valid_groups:
            inv_num = int(valid_groups[-1])
        elif digit_groups:
            inv_num = int(digit_groups[-1])

    if not inv_num:
        return None

    # 2. Date (Ngày tháng dd/mm/yyyy)
    date_str = None
    m_date1 = re.search(r'(\d{1,2})/(\d{1,2})/(\d{4})', full_text)
    if m_date1:
        d, m, y = m_date1.groups()
        date_str = f"{int(d):02d}/{int(m):02d}/{int(y):04d}"
    else:
        m_date2 = re.search(r'Ng[àa]y[^\d]*(\d{1,2})[^\d]+th[áa]ng[^\d]*(\d{1,2})[^\d]+n[ăa]m[^\d]*(\d{4})', full_text, re.IGNORECASE)
        if m_date2:
            d, m, y = m_date2.groups()
            date_str = f"{int(d):02d}/{int(m):02d}/{int(y):04d}"

    # 3. Quantity (Số lượng)
    # On Vietnamese invoices: comma = decimal separator, dot = thousands separator
    # So "26,918" = 26.918 (quantity), "22.290" = 22290 (unit price), "600.000" = 600000 (amount)
    # Strategy: Find the first standalone comma-decimal number (XX,XXX) in the data
    # area after the table headers. This is always the quantity value.
    qty = None

    # Find where the table data starts (after the last header line like "Thanh tien"/"Amount")
    data_start = 0
    for i, line in enumerate(lines):
        if re.search(r'Thanh\s*ti[êeề]n|^\(Amount\)$|^\(Thanh\s*ti[êeề]n\)$', line.strip(), re.IGNORECASE):
            data_start = i + 1

    # Search for the first standalone comma-decimal number in the data area
    # Comma-decimal format: "26,918" or "27,791" or "22,840" (digits,digits)
    # This distinguishes quantity from money amounts which use dots (600.000, 22.290)
    for j in range(data_start, len(lines)):
        stripped = lines[j].strip()
        m_num = re.match(r'^(\d{1,5},\d{1,3})$', stripped)
        if m_num:
            raw = m_num.group(1)
            try:
                # Comma is decimal separator: "26,918" -> 26.918
                qty = float(raw.replace(',', '.'))
                break
            except ValueError:
                pass

    # 4. Money Amounts (Tiền hàng, Thuế GTGT, Tổng tiền)
    subtotal = None
    tax = 0
    total = None

    # Subtotal (Tiền trước thuế)
    m_subtotal = re.search(r'(?:C[ộo]ng ti[ềe]n h[àa]ng|T[ổo]ng ti[ềe]n ch[ưua] thu[ếe]|Ti[ềe]n ch[ưua] thu[ếe]|Total before VAT)[^\d\n]*\n?(?:\([^\)]*\):?\s*\n?)?([0-9.,\']+)', full_text, re.IGNORECASE)
    if m_subtotal:
        try:
            subtotal = int(m_subtotal.group(1).replace('.', '').replace(',', '').replace("'", ""))
        except ValueError:
            pass

    # Tax (Tiền thuế GTGT)
    m_tax = re.search(r'(?:Ti[ềe]n thu[ếe]\s*GTGT|T[iêếe]n thu[ếe]|VAT amount)[^\d\n]*(?:\d+(?:[.,]\d+)?\s*%?[^\d\n]*)?\n?\s*([0-9.,\']+)', full_text, re.IGNORECASE)
    if m_tax:
        try:
            tax = int(m_tax.group(1).replace('.', '').replace(',', '').replace("'", ""))
        except ValueError:
            tax = 0

    # Total (Tổng tiền thanh toán)
    m_total = re.search(r'(?:T[oổ]ng s[ốseóo] ti[ềe]n thanh to[áa]n|Gi[áa] tr[ịi] thanh to[áa]n|T[ổo]ng ti[ềe]n thanh to[áa]n|Total amount)[^\d\n]*\n?([0-9.,\']+)', full_text, re.IGNORECASE)
    if m_total:
        try:
            total = int(m_total.group(1).replace('.', '').replace(',', '').replace("'", ""))
        except ValueError:
            pass

    if subtotal is None and total is not None:
        subtotal = total - tax
    if total is None and subtotal is not None:
        total = subtotal + tax

    # Auto-correct subtotal if math doesn't match
    if subtotal is not None and total is not None and tax is not None:
        if subtotal + tax != total:
            subtotal = total - tax

    return {
        'inv_num': inv_num,
        'date_str': date_str,
        'quantity': qty,
        'subtotal': subtotal,
        'tax': tax,
        'total': total
    }

def format_cell_row(sheet, row_idx):
    """Apply center alignment and number formatting (thousands separator) to a row."""
    for c in range(1, 7):
        sheet.cell(row=row_idx, column=c).alignment = CENTER_ALIGN
    
    # Qty
    c_qty = sheet.cell(row=row_idx, column=3)
    if isinstance(c_qty.value, (int, float)):
        c_qty.number_format = FMT_QTY
        
    # Money
    sheet.cell(row=row_idx, column=4).number_format = FMT_CURRENCY
    sheet.cell(row=row_idx, column=5).number_format = FMT_CURRENCY
    sheet.cell(row=row_idx, column=6).number_format = FMT_CURRENCY

def apply_old_style(sheet, row_idx):
    """Mark old data row with red text and clear any custom background."""
    for c in range(1, 7):
        cell = sheet.cell(row=row_idx, column=c)
        cell.font = FONT_RED
        cell.fill = NO_FILL  # Reset background to table default

def apply_new_style(sheet, row_idx):
    """New data row: default font, clear any custom background."""
    for c in range(1, 7):
        cell = sheet.cell(row=row_idx, column=c)
        cell.font = Font()  # Default font (black)
        cell.fill = NO_FILL  # Reset background to table default

def update_table_ranges(sheet):
    """Expand Excel Table objects (ListObjects) to encompass all data rows."""
    last_row = sheet.max_row
    while last_row > 1:
        vals = [sheet.cell(last_row, c).value for c in range(1, 7)]
        if any(v is not None for v in vals):
            break
        last_row -= 1

    for tbl_name in list(sheet.tables):
        tbl = sheet.tables[tbl_name]
        if hasattr(tbl, 'ref') and ":" in tbl.ref:
            start_cell, end_cell = tbl.ref.split(":")
            m = re.match(r'([A-Za-z]+)(\d+)', end_cell)
            if m:
                col_letter = m.group(1)
                tbl.ref = f"{start_cell}:{col_letter}{last_row}"
                print(f"Đã mở rộng Excel Table '{tbl_name}' thành vùng: {tbl.ref}")

def main():
    print("=" * 60)
    print("      HỆ THỐNG XỬ LÝ HÓA ĐƠN TỰ ĐỘNG (PDF -> EXCEL TABLE)")
    print("=" * 60)

    if not os.path.exists(EXCEL_PATH):
        print(f"LỖI: Không tìm thấy file Excel tại {EXCEL_PATH}")
        return

    try:
        wb = openpyxl.load_workbook(EXCEL_PATH)
    except PermissionError:
        print("\nLỖI: Tệp Xăng.xlsx đang mở trên ứng dụng Excel.")
        print("Vui lòng ĐÓNG tệp Xăng.xlsx trên màn hình rồi chạy lại run.bat!")
        return

    sheet = wb['Sheet1']

    # Ensure all existing data rows are formatted properly + mark as OLD (black bg, white text)
    for r in range(2, sheet.max_row + 1):
        if sheet.cell(r, 1).value is not None:
            format_cell_row(sheet, r)
            apply_old_style(sheet, r)

    # Existing invoices
    existing_invoices = set()
    # Track rows with missing quantity for updating
    missing_qty_rows = {}  # inv_num -> row_idx
    for r in range(2, sheet.max_row + 1):
        v = sheet.cell(r, 1).value
        if v is not None:
            v_str = str(v).strip()
            m = re.search(r'(\d+)', v_str)
            if m:
                inv_num_val = int(m.group(1))
                existing_invoices.add(inv_num_val)
                # Check if quantity (column C) is missing
                qty_val = sheet.cell(r, 3).value
                if qty_val is None or qty_val == '' or qty_val == 0:
                    missing_qty_rows[inv_num_val] = r

    print(f"Hóa đơn hiện có trong Excel: {sorted(list(existing_invoices))}")
    if missing_qty_rows:
        print(f"Hóa đơn thiếu Số lượng cần cập nhật: {sorted(list(missing_qty_rows.keys()))}")

    if not os.path.exists(PDF_DIR):
        print(f"LỖI: Không tìm thấy thư mục Hóa đơn tại {PDF_DIR}")
        return

    pdf_files = [os.path.join(PDF_DIR, f) for f in os.listdir(PDF_DIR) if f.lower().endswith('.pdf')]
    print(f"Tìm thấy {len(pdf_files)} tệp PDF trong thư mục Hóa đơn.\n")

    new_invoices = []
    updated_qty_count = 0
    total_files = len(pdf_files)
    for idx, pdf_path in enumerate(pdf_files, 1):
        filename = os.path.basename(pdf_path)
        print(f"[{idx}/{total_files}] Đang đọc và nhận diện: {filename}...", end="", flush=True)
        info = parse_pdf_invoice(pdf_path)
        if not info:
            print(" -> BỎ QUA (Không trích xuất được thông tin)", flush=True)
            continue
            
        inv_num = info['inv_num']
        if inv_num in existing_invoices:
            # Check if we need to update missing quantity
            if inv_num in missing_qty_rows and info['quantity'] is not None:
                row_idx = missing_qty_rows[inv_num]
                sheet.cell(row=row_idx, column=3, value=info['quantity'])
                format_cell_row(sheet, row_idx)
                print(f" -> CẬP NHẬT SỐ LƯỢNG: {info['quantity']:.3f} (Dòng {row_idx})", flush=True)
                del missing_qty_rows[inv_num]
                updated_qty_count += 1
            else:
                print(f" -> ĐÃ CÓ TRONG EXCEL (Số HĐ: {inv_num})", flush=True)
            continue

        total_display = f"{info['total']:,}" if isinstance(info['total'], (int, float)) else "0"
        print(f" -> OK (Số HĐ: {inv_num} | Tổng: {total_display} VNĐ)", flush=True)
        new_invoices.append(info)
        existing_invoices.add(inv_num)

    if not new_invoices:
        update_table_ranges(sheet)
        try:
            wb.save(EXCEL_PATH)
        except PermissionError:
            print("\nLỖI GHI FILE: Tệp Xăng.xlsx đang được mở trong Excel. Vui lòng đóng Excel và chạy lại run.bat!")
            return
        if updated_qty_count > 0:
            print(f"\n=> KẾT QUẢ: Đã cập nhật Số lượng cho {updated_qty_count} hóa đơn. Không có hóa đơn mới.")
        else:
            print("\n=> KẾT QUẢ: Tất cả các hóa đơn PDF đều đã nằm trong bảng Excel Table.")
        return

    new_invoices.sort(key=lambda x: x['inv_num'])

    next_row = 2
    while next_row <= sheet.max_row:
        if sheet.cell(next_row, 1).value is None:
            break
        next_row += 1

    print("\n--- ĐANG GHI HÓA ĐƠN MỚI VÀO EXCEL TABLE ---")
    for item in new_invoices:
        sheet.cell(row=next_row, column=1, value=item['inv_num'])
        sheet.cell(row=next_row, column=2, value=item['date_str'])
        sheet.cell(row=next_row, column=3, value=item['quantity'])
        sheet.cell(row=next_row, column=4, value=item['subtotal'])
        sheet.cell(row=next_row, column=5, value=item['tax'])
        sheet.cell(row=next_row, column=6, value=item['total'])
        
        # Apply alignment, number format and red highlight on col A for new row
        format_cell_row(sheet, next_row)
        apply_new_style(sheet, next_row)
        
        total_display = f"{item['total']:,}" if isinstance(item['total'], (int, float)) else "0"
        print(f" + Đã thêm vào Table: HĐ {item['inv_num']} | Ngày: {item['date_str']} | SL: {item['quantity']} | Tổng: {total_display} VNĐ")
        next_row += 1

    update_table_ranges(sheet)

    try:
        wb.save(EXCEL_PATH)
        print(f"\n===> THÀNH CÔNG: Đã thêm mới {len(new_invoices)} hóa đơn vào bảng Excel Table!")
    except PermissionError:
        print("\nLỖI GHI FILE: Tệp Xăng.xlsx đang được mở trong Excel. Vui lòng đóng Excel và chạy lại run.bat!")

if __name__ == '__main__':
    main()
