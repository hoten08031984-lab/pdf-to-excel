import os
import sys
import re
import datetime
import fitz  # PyMuPDF
from rapidocr_onnxruntime import RapidOCR
import openpyxl
from openpyxl.styles import Alignment

sys.stdout.reconfigure(encoding='utf-8')

ocr = RapidOCR()

BASE_DIR = os.path.dirname(os.path.abspath(__file__))
EXCEL_PATH = os.path.join(BASE_DIR, 'Xăng.xlsx')
PDF_DIR = os.path.join(BASE_DIR, 'Hóa đơn')

# Styles
CENTER_ALIGN = Alignment(horizontal='center', vertical='center')
FMT_CURRENCY = '_-* #,##0_-;\\-* #,##0_-;_-* "-"_-;_-@_-'
FMT_QTY = '#,##0.000'

def parse_pdf_invoice(pdf_path):
    """Extract invoice details using PyMuPDF + RapidOCR."""
    doc = fitz.open(pdf_path)
    page = doc[0]
    pix = page.get_pixmap(dpi=200)
    img_bytes = pix.tobytes('png')
    
    result, _ = ocr(img_bytes)
    if not result:
        return None
        
    lines = [item[1].strip() for item in result]
    full_text = "\n".join(lines)
    
    # 1. Invoice Number (Số hóa đơn)
    inv_num = None
    # Exclude lines containing bank accounts or tax codes
    clean_text = "\n".join([line for line in full_text.split("\n") if not re.search(r'Acc\s*No|t[àa]i khoản|Tax\s*Code|mã số thuế', line, re.IGNORECASE)])
    
    m_inv = re.search(r'\(No\.?\):?\s*(\d{1,8})', clean_text, re.IGNORECASE)
    if not m_inv:
        m_inv = re.search(r'(?:S[ốóe60ô]|So|No|S6|S0)[:\s\.]*(?:\(?No\.?\)?[:\s\.]*)?\n?\s*(\d{1,8})', clean_text, re.IGNORECASE)
    if not m_inv:
        m_inv = re.search(r'(?:S[ốóe60ô]|So|S6|S0)[:.\s]+\n?\s*(\d{1,8})', clean_text, re.IGNORECASE)
    
    if m_inv:
        inv_num = int(m_inv.group(1))
    else:
        # Fallback: find all digit groups in filename, pick the last group with length >= 4 (e.g. 164323 in HDGTGT_K26TXT_164323.pdf)
        digit_groups = re.findall(r'\d+', os.path.basename(pdf_path))
        valid_groups = [g for g in digit_groups if len(g) >= 4]
        if valid_groups:
            inv_num = int(valid_groups[-1])
        elif digit_groups:
            inv_num = int(digit_groups[-1])

    if not inv_num:
        return None

    # 2. Date (Ngày tháng dd/mm/yyyy)
    date_str = None
    m_date1 = re.search(r'(\d{1,2})/(\d{1,2})/(\d{4})', full_text)
    if m_date1:
        d, m, y = m_date1.groups()
        date_str = f"{int(d):02d}/{int(m):02d}/{int(y):04d}"
    else:
        m_date2 = re.search(r'Ng[àa]y[^\d]*(\d{1,2})[^\d]+th[áa]ng[^\d]*(\d{1,2})[^\d]+n[ăa]m[^\d]*(\d{4})', full_text, re.IGNORECASE)
        if m_date2:
            d, m, y = m_date2.groups()
            date_str = f"{int(d):02d}/{int(m):02d}/{int(y):04d}"

    # 3. Quantity (Số lượng)
    qty = None
    m_qty = re.search(r'(?:Lit|D[ầa]u|h[ộo]p|b[ìi]nh|c[áa]i|b[ộo]|chi[ếe]c|kg|m[ée]t|g[óo]i|lon|chai|th[ùu]ng|l[ôo])[^\n\d]*([0-9.,]+)', full_text, re.IGNORECASE)
    if m_qty:
        try:
            val_str = m_qty.group(1).replace('.', '').replace(',', '.') if m_qty.group(1).count('.') > 1 or (',' in m_qty.group(1) and '.' in m_qty.group(1) and m_qty.group(1).find('.') < m_qty.group(1).find(',')) else m_qty.group(1).replace(',', '.')
            qty = float(val_str)
        except ValueError:
            pass

    if qty is None:
        matches = re.findall(r'(\d+[\.,]\d{2,3})', full_text)
        if matches:
            try:
                qty = float(matches[0].replace(',', '.'))
            except ValueError:
                pass

    # 4. Money Amounts (Tiền hàng, Thuế GTGT, Tổng tiền)
    subtotal = None
    tax = 0
    total = None

    # Subtotal (Tiền trước thuế)
    m_subtotal = re.search(r'(?:C[ộo]ng ti[ềe]n h[àa]ng|T[ổo]ng ti[ềe]n ch[ưua] thu[ếe]|Ti[ềe]n ch[ưua] thu[ếe]|Total before VAT)[:\s]*\n?([0-9.,]+)', full_text, re.IGNORECASE)
    if m_subtotal:
        try:
            subtotal = int(m_subtotal.group(1).replace('.', '').replace(',', ''))
        except ValueError:
            pass

    # Tax (Tiền thuế GTGT)
    m_tax = re.search(r'(?:T[iêếe]n thue\s*GTGT|T[iêếe]n thu[ếe]|VAT amount)[:\s]*\n?([0-9.,]+)', full_text, re.IGNORECASE)
    if m_tax:
        try:
            tax = int(m_tax.group(1).replace('.', '').replace(',', ''))
        except ValueError:
            tax = 0

    # Total (Tổng tiền thanh toán)
    m_total = re.search(r'(?:Tong s[ốseó] ti[ềe]n thanh to[áa]n|Gi[áa] tr[ịi] thanh to[áa]n|T[ổo]ng ti[ềe]n thanh to[áa]n|Total amount|T[ổo]ng c[ộo]ng ti[ềe]n thanh to[áa]n)[:\s]*\n?([0-9.,]+)', full_text, re.IGNORECASE)
    if m_total:
        try:
            total = int(m_total.group(1).replace('.', '').replace(',', ''))
        except ValueError:
            pass

    if subtotal is None and total is not None:
        subtotal = total - tax
    if total is None and subtotal is not None:
        total = subtotal + tax

    return {
        'inv_num': inv_num,
        'date_str': date_str,
        'quantity': qty,
        'subtotal': subtotal,
        'tax': tax,
        'total': total
    }

def format_cell_row(sheet, row_idx):
    """Apply center alignment and number formatting (thousands separator) to a row."""
    for c in range(1, 7):
        sheet.cell(row=row_idx, column=c).alignment = CENTER_ALIGN
    
    # Qty
    c_qty = sheet.cell(row=row_idx, column=3)
    if isinstance(c_qty.value, (int, float)):
        c_qty.number_format = FMT_QTY
        
    # Money
    sheet.cell(row=row_idx, column=4).number_format = FMT_CURRENCY
    sheet.cell(row=row_idx, column=5).number_format = FMT_CURRENCY
    sheet.cell(row=row_idx, column=6).number_format = FMT_CURRENCY

def update_table_ranges(sheet):
    """Expand Excel Table objects (ListObjects) to encompass all data rows."""
    last_row = sheet.max_row
    while last_row > 1:
        vals = [sheet.cell(last_row, c).value for c in range(1, 7)]
        if any(v is not None for v in vals):
            break
        last_row -= 1

    for tbl_name in list(sheet.tables):
        tbl = sheet.tables[tbl_name]
        if hasattr(tbl, 'ref') and ":" in tbl.ref:
            start_cell, end_cell = tbl.ref.split(":")
            m = re.match(r'([A-Za-z]+)(\d+)', end_cell)
            if m:
                col_letter = m.group(1)
                tbl.ref = f"{start_cell}:{col_letter}{last_row}"
                print(f"Đã mở rộng Excel Table '{tbl_name}' thành vùng: {tbl.ref}")

def main():
    print("=" * 60)
    print("      HỆ THỐNG XỬ LÝ HÓA ĐƠN TỰ ĐỘNG (PDF -> EXCEL TABLE)")
    print("=" * 60)

    if not os.path.exists(EXCEL_PATH):
        print(f"LỖI: Không tìm thấy file Excel tại {EXCEL_PATH}")
        return

    try:
        wb = openpyxl.load_workbook(EXCEL_PATH)
    except PermissionError:
        print("\nLỖI: Tệp Xăng.xlsx đang mở trên ứng dụng Excel.")
        print("Vui lòng ĐÓNG tệp Xăng.xlsx trên màn hình rồi chạy lại run.bat!")
        return

    sheet = wb['Sheet1']

    # Ensure all existing data rows are formatted properly
    for r in range(2, sheet.max_row + 1):
        if sheet.cell(r, 1).value is not None:
            format_cell_row(sheet, r)

    # Existing invoices
    existing_invoices = set()
    for r in range(2, sheet.max_row + 1):
        v = sheet.cell(r, 1).value
        if v is not None:
            v_str = str(v).strip()
            m = re.search(r'(\d+)', v_str)
            if m:
                existing_invoices.add(int(m.group(1)))

    print(f"Hóa đơn hiện có trong Excel: {sorted(list(existing_invoices))}")

    if not os.path.exists(PDF_DIR):
        print(f"LỖI: Không tìm thấy thư mục Hóa đơn tại {PDF_DIR}")
        return

    pdf_files = [os.path.join(PDF_DIR, f) for f in os.listdir(PDF_DIR) if f.lower().endswith('.pdf')]
    print(f"Tìm thấy {len(pdf_files)} tệp PDF trong thư mục Hóa đơn.\n")

    new_invoices = []
    for pdf_path in pdf_files:
        filename = os.path.basename(pdf_path)
        info = parse_pdf_invoice(pdf_path)
        if not info:
            print(f"-> Bỏ qua {filename}: Không trích xuất được thông tin.")
            continue
            
        inv_num = info['inv_num']
        if inv_num in existing_invoices:
            print(f"-> Bỏ qua HĐ {inv_num} ({filename}): Đã có trong Excel Table.")
            continue

        new_invoices.append(info)
        existing_invoices.add(inv_num)

    if not new_invoices:
        update_table_ranges(sheet)
        wb.save(EXCEL_PATH)
        print("\n=> KẾT QUẢ: Tất cả các hóa đơn PDF đều đã nằm trong bảng Excel Table.")
        return

    new_invoices.sort(key=lambda x: x['inv_num'])

    next_row = 2
    while next_row <= sheet.max_row:
        if sheet.cell(next_row, 1).value is None:
            break
        next_row += 1

    print("\n--- ĐANG GHI HÓA ĐƠN MỚI VÀO EXCEL TABLE ---")
    for item in new_invoices:
        sheet.cell(row=next_row, column=1, value=item['inv_num'])
        sheet.cell(row=next_row, column=2, value=item['date_str'])
        sheet.cell(row=next_row, column=3, value=item['quantity'])
        sheet.cell(row=next_row, column=4, value=item['subtotal'])
        sheet.cell(row=next_row, column=5, value=item['tax'])
        sheet.cell(row=next_row, column=6, value=item['total'])
        
        # Apply alignment and number format to new row
        format_cell_row(sheet, next_row)
        
        total_display = f"{item['total']:,}" if isinstance(item['total'], (int, float)) else "0"
        print(f" + Đã thêm vào Table: HĐ {item['inv_num']} | Ngày: {item['date_str']} | SL: {item['quantity']} | Tổng: {total_display} VNĐ")
        next_row += 1

    update_table_ranges(sheet)

    try:
        wb.save(EXCEL_PATH)
        print(f"\n===> THÀNH CÔNG: Đã thêm mới {len(new_invoices)} hóa đơn vào bảng Excel Table!")
    except PermissionError:
        print("\nLỖI GHI FILE: Tệp Xăng.xlsx đang được mở trong Excel. Vui lòng đóng Excel và chạy lại run.bat!")

if __name__ == '__main__':
    main()
